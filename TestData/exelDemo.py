import openpyxl

book = openpyxl.load_workbook("/Users/natulik/PycharmProjects/PythonSelFramew/TestData/PythonDemo2.xlsx")

sheet = book.active

Dict = {}

# cell = sheet.cell(row=13, column=4)

# print(cell.value)

# newvalue = sheet.cell(row=5, column=5).value = "new value"

# sheet.delete_rows(5, 23)
# sheet.delete_cols(5, 6)
# print(newvalue)
# print(sheet.max_row)
# book.save("/Users/natulik/PycharmProjects/PythonSelFramew/TestData/PythonDemo2.xlsx")
# print(sheet.max_column)
# print(sheet['C3'].value)

# for i in range(1, sheet.max_row+1):
#     for j in range(1, sheet.max_column+1):
#         print(sheet.cell(row=i, column=j).value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):
            # Dict["firstname"] = "Nata"
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)