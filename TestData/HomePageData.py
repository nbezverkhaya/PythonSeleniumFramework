import openpyxl

class HomePageData:
    test_HomePage_data = [{"firstname":"Nata", "email":"natata@gmail.com", "password":"123321", "gender":"Female"}, {"firstname":"Tata", "email":"tata@gmail.com", "password":"432432", "gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("/Users/natulik/PycharmProjects/PythonSelFramew/TestData/PythonDemo2.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    # Dict["firstname"] = "Natalia"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]